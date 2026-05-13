"""Parser genérico de Excel: lee un archivo y devuelve filas dict[col->valor].

Hace UNA cosa: validar headers + leer filas. No conoce las reglas de negocio.
"""
from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import load_workbook


class ExcelParseError(Exception):
    pass


def parse_workbook(content: bytes, expected_columns: list[str]) -> list[dict]:
    """Lee el primer worksheet del Excel y devuelve una lista de dicts.

    - Valida que los headers de la primera fila coincidan EXACTAMENTE con
      `expected_columns` (orden flexible, contenido estricto).
    - Filas vacías se ignoran.
    - Lanza ExcelParseError si los headers no empatan.
    """
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise ExcelParseError(f"No se pudo leer el archivo Excel: {e}") from e

    ws = wb.active
    if ws is None:
        raise ExcelParseError("El archivo no tiene hojas.")

    rows: Iterable = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise ExcelParseError("El archivo está vacío.")

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    headers = [h for h in headers if h]  # quita columnas vacías al final

    missing = [c for c in expected_columns if c not in headers]
    extra = [c for c in headers if c not in expected_columns]
    if missing or extra:
        msg = []
        if missing:
            msg.append(f"Faltan columnas: {', '.join(missing)}")
        if extra:
            msg.append(f"Columnas no esperadas: {', '.join(extra)}")
        raise ExcelParseError(" | ".join(msg))

    out: list[dict] = []
    for row in rows:
        if row is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        record: dict = {}
        for i, col in enumerate(headers):
            record[col] = row[i] if i < len(row) else None
        out.append(record)

    return out
