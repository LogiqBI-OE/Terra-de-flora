"""Genera los 4 templates Excel descargables.

Cada template tiene:
  - Una hoja con el header (columnas obligatorias)
  - 1-2 filas de ejemplo para que el usuario vea el formato esperado

Para cambiar las columnas de un template, edita TEMPLATES abajo.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ┌─────────────────────────────────────────────────────────────────────────────┐
# │ DEFINICIÓN DE TEMPLATES                                                     │
# │ Para cambiar columnas: edita aquí. El parser leerá los mismos headers.      │
# └─────────────────────────────────────────────────────────────────────────────┘
TEMPLATES: dict[str, dict] = {
    "inventario": {
        "columnas": ["SKU", "Producto", "Planta", "Fecha", "Cantidad", "Unidad"],
        "ejemplos": [
            ["AGU-001", "Aguacate Hass A", "MTY", "2025-05-12", 5000, "kg"],
            ["ACT-010", "Aceite Aguacate 1L", "GDL", "2025-05-12", 1200, "L"],
        ],
    },
    "produccion": {
        "columnas": ["SKU", "Producto", "Planta", "Fecha", "Cantidad"],
        "ejemplos": [
            ["ACT-010", "Aceite Aguacate 1L", "MTY", "2025-05-19", 800],
        ],
    },
    "compras": {
        "columnas": ["SKU", "Producto", "Planta", "OC", "ETA", "Cantidad"],
        "ejemplos": [
            ["AGU-001", "Aguacate Hass A", "MTY", "OC-2025-115", "2025-05-26", 3000],
        ],
    },
    "demanda": {
        "columnas": ["SKU", "Producto", "Cliente", "Fecha", "Cantidad"],
        "ejemplos": [
            ["ACT-010", "Aceite Aguacate 1L", "CHEDRAUI", "2025-05-19", 500],
            ["ACT-010", "Aceite Aguacate 1L", "WALMART", "2025-05-19", 700],
        ],
    },
}


def generar_template(tipo: str) -> bytes:
    """Devuelve los bytes de un .xlsx con el template del tipo solicitado."""
    if tipo not in TEMPLATES:
        raise ValueError(f"Tipo desconocido: {tipo}")

    spec = TEMPLATES[tipo]
    wb = Workbook()
    ws = wb.active
    ws.title = tipo.capitalize()

    # Header
    header_fill = PatternFill("solid", fgColor="A8D060")
    header_font = Font(bold=True, color="0B0F08")
    for col_idx, col in enumerate(spec["columnas"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Ejemplos
    for row_idx, ejemplo in enumerate(spec["ejemplos"], start=2):
        for col_idx, value in enumerate(ejemplo, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto width sencillo
    for col_idx, col in enumerate(spec["columnas"], start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = max(14, len(col) + 2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
